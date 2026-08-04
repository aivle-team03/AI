from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any

DEFAULT_TEMPLATE_PATH = Path(
    r"C:\Users\User\Desktop\KT에이블\빅프로젝트\청취조사에_의한_유해위험요인_가이드라인.xlsx"
)
DEFAULT_OUTPUT_DIR = Path("output/worker_feedback_reports")

ROW_CELL_MAP = {
    "category": "C3",
    "risk": "C18",
    "board_created_at": "C5",
    "category_name": "C8",
    "board_contents": "A11",
    "status": "H7",
    "board_image_url": "A21",
    "location": "C7",
    "completed_at": "C33",
    "handler_name": "H28",
    "content": "C30",
    "image_url": "A36",
}


def _load_workbook(template_path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to write xlsx files. "
            "Install it with: pip install openpyxl"
        ) from exc

    return load_workbook(template_path)


def _row_to_dict(row: Any) -> dict[str, Any]:
    if hasattr(row, "model_dump"):
        return row.model_dump(mode="json")
    return dict(row or {})


def _safe_filename_part(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r'[<>:"/\\|?*]', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text[:40] or "worker_feedback"


def _set_cell_value(worksheet, cell: str, value: Any) -> None:
    if value is None:
        value = ""
    worksheet[cell] = value


def _border_with_side(border, *, left=None, right=None, top=None, bottom=None):
    from openpyxl.styles import Border

    return Border(
        left=left if left is not None else border.left,
        right=right if right is not None else border.right,
        top=top if top is not None else border.top,
        bottom=bottom if bottom is not None else border.bottom,
        diagonal=border.diagonal,
        diagonal_direction=border.diagonal_direction,
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        vertical=border.vertical,
        horizontal=border.horizontal,
    )


def _restore_merged_cell_borders(worksheet) -> None:
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        source_border = copy(worksheet.cell(min_row, min_col).border)

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = worksheet.cell(row, col)
                border = copy(cell.border)
                cell.border = _border_with_side(
                    border,
                    left=source_border.left if col == min_col else None,
                    right=source_border.right if col == max_col else None,
                    top=source_border.top if row == min_row else None,
                    bottom=source_border.bottom if row == max_row else None,
                )


def fill_worker_feedback_excel_report(
    row: Any,
    output_path: Path | str,
    template_path: Path | str = DEFAULT_TEMPLATE_PATH,
    sheet_name: str | None = None,
) -> str:
    template_path = Path(template_path)
    output_path = Path(output_path)
    row_data = _row_to_dict(row)

    workbook = _load_workbook(template_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    for source_field, cell in ROW_CELL_MAP.items():
        _set_cell_value(worksheet, cell, row_data.get(source_field))

    _restore_merged_cell_borders(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return str(output_path)


def fill_worker_feedback_excel_reports(
    corrected_rows: list[Any],
    template_path: Path | str = DEFAULT_TEMPLATE_PATH,
    output_dir: Path | str = DEFAULT_OUTPUT_DIR,
    sheet_name: str | None = None,
) -> list[str]:
    output_dir = Path(output_dir)
    output_paths: list[str] = []

    for index, row in enumerate(corrected_rows, start=1):
        row_data = _row_to_dict(row)
        name_part = _safe_filename_part(
            row_data.get("category_name")
            or row_data.get("category")
            or row_data.get("location")
        )
        output_path = output_dir / f"worker_feedback_{index:03d}_{name_part}.xlsx"
        output_paths.append(
            fill_worker_feedback_excel_report(
                row,
                output_path=output_path,
                template_path=template_path,
                sheet_name=sheet_name,
            )
        )

    return output_paths
